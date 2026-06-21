import csv
import json
from pathlib import Path
from typing import Any

from app.connectors.base import BaseConnector, ColumnMetadata, ProfileConfig, TableMetadata


class FileConnector(BaseConnector):
    def __init__(self, source_name: str, system_type: str, secret_reference: str | None = None) -> None:
        super().__init__(source_name, secret_reference)
        self.system_type = system_type.lower()
        self.file_path = _resolve_file_path(secret_reference)

    def test_connection(self) -> bool:
        return self.file_path.is_file()

    def list_schemas(self) -> list[str]:
        return ["uploaded"]

    def list_tables(self, scan_mode: str = "metadata_only") -> list[TableMetadata]:
        rows, columns = self._read_rows()
        return [
            TableMetadata(
                database_name=self.file_path.stem,
                schema_name="uploaded",
                table_name=self.file_path.stem,
                row_count=len(rows),
                row_count_is_estimated=False,
                columns=[ColumnMetadata(column, _infer_type([row.get(column) for row in rows]), True, False, False, None) for column in columns],
            )
        ]

    def profile_columns(self, schema_name: str, table_name: str, columns: list[str], row_count: int, config: ProfileConfig) -> dict[str, dict[str, Any]]:
        rows, _columns = self._read_rows()
        rows = rows[: config.row_limit]
        profiles: dict[str, dict[str, Any]] = {}
        for column in columns[: config.max_columns]:
            values = [row.get(column) for row in rows]
            null_count = sum(1 for value in values if value in (None, ""))
            distinct_values = {str(value) for value in values if value not in (None, "")}
            profiles[column] = {
                "null_percentage": round((null_count / len(rows)) * 100, 2) if rows else 0,
                "distinct_count": len(distinct_values),
                "sampled_row_count": len(rows),
            }
        return profiles

    def find_duplicate_rows(self, schema_name: str, table_name: str, columns: list[str], config: ProfileConfig) -> dict[str, Any]:
        rows, _columns = self._read_rows()
        rows = rows[: config.row_limit]
        selected_columns = columns[: config.max_columns]
        duplicate_rows = _count_duplicate_rows(rows, selected_columns)
        return {
            "duplicate_rows": duplicate_rows,
            "sampled_row_count": len(rows),
            "columns_checked": selected_columns,
        }

    def _read_rows(self) -> tuple[list[dict[str, Any]], list[str]]:
        if self.system_type == "csv":
            return _read_csv(self.file_path)
        if self.system_type == "json":
            return _read_json(self.file_path)
        if self.system_type == "excel":
            return _read_excel(self.file_path)
        if self.system_type == "parquet":
            return _read_parquet(self.file_path)
        raise ValueError(f"Unsupported file source type: {self.system_type}")


def _resolve_file_path(secret_reference: str | None) -> Path:
    if not secret_reference or not secret_reference.startswith("file://"):
        raise ValueError("Uploaded file sources require a file:// reference")
    return Path(secret_reference.removeprefix("file://"))


def _read_csv(file_path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    with file_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        rows = [dict(row) for row in reader]
        return rows, list(reader.fieldnames or [])


def _read_json(file_path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    data = json.loads(file_path.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        records = data.get("records") or data.get("data") or [data]
    else:
        records = data
    if not isinstance(records, list) or any(not isinstance(row, dict) for row in records):
        raise ValueError("JSON source files must contain an object or an array of objects")
    rows = [dict(row) for row in records]
    columns = sorted({key for row in rows for key in row.keys()})
    return rows, columns


def _read_excel(file_path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(iterator, [])]
    columns = [header or f"column_{index + 1}" for index, header in enumerate(headers)]
    rows = [{columns[index]: value for index, value in enumerate(values)} for values in iterator]
    return rows, columns


def _read_parquet(file_path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    try:
        import pyarrow.parquet as pq
    except ImportError as exc:
        raise ValueError("Install pyarrow to scan Parquet source files") from exc

    table = pq.read_table(file_path)
    rows = table.to_pylist()
    return rows, list(table.column_names)


def _count_duplicate_rows(rows: list[dict[str, Any]], columns: list[str]) -> int:
    if not rows or not columns:
        return 0
    seen: dict[tuple[str, ...], int] = {}
    duplicate_rows = 0
    for row in rows:
        signature = tuple(_normalized_cell(row.get(column)) for column in columns)
        if all(value == "" for value in signature):
            continue
        previous_count = seen.get(signature, 0)
        if previous_count:
            duplicate_rows += 1
        seen[signature] = previous_count + 1
    return duplicate_rows


def _normalized_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().casefold()


def _infer_type(values: list[Any]) -> str:
    non_empty = [value for value in values if value not in (None, "")]
    if not non_empty:
        return "unknown"
    if all(isinstance(value, bool) or str(value).lower() in {"true", "false"} for value in non_empty):
        return "boolean"
    if all(_is_int(value) for value in non_empty):
        return "integer"
    if all(_is_float(value) for value in non_empty):
        return "decimal"
    return "string"


def _is_int(value: Any) -> bool:
    try:
        int(value)
    except (TypeError, ValueError):
        return False
    return True


def _is_float(value: Any) -> bool:
    try:
        float(value)
    except (TypeError, ValueError):
        return False
    return True
