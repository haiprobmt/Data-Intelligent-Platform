import csv
import zipfile
from pathlib import Path
from xml.etree import ElementTree

from app.services.ocr import get_ocr_provider


def extract_text_from_file(file_path: str | None) -> tuple[str, dict]:
    if not file_path:
        return "", {"method": "missing_file_path", "pages": []}
    path = Path(file_path)
    if not path.exists():
        return "", {"method": "missing_file", "pages": []}

    extension = path.suffix.lower()
    if extension in {".txt", ".md"}:
        return path.read_text(encoding="utf-8", errors="ignore"), {"method": "plain_text", "pages": []}
    if extension == ".csv":
        return _extract_csv(path), {"method": "csv", "pages": []}
    if extension in {".xlsx", ".xls"}:
        return _extract_excel(path), {"method": "openpyxl", "pages": []}
    if extension == ".docx":
        return _extract_docx(path), {"method": "docx_xml", "pages": []}
    if extension == ".pdf":
        return _extract_pdf(path)
    return "", {"method": "unsupported", "pages": []}


def _extract_csv(path: Path) -> str:
    with path.open(newline="", encoding="utf-8-sig", errors="ignore") as handle:
        reader = csv.reader(handle)
        return "\n".join(" | ".join(str(value) for value in row) for row in reader)


def _extract_excel(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value) for value in row]
            if any(values):
                lines.append(" | ".join(values))
    return "\n".join(lines)


def _extract_docx(path: Path) -> str:
    namespaces = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml")
    root = ElementTree.fromstring(xml)
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", namespaces):
        text = "".join(node.text or "" for node in paragraph.findall(".//w:t", namespaces)).strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def _extract_pdf(path: Path) -> tuple[str, dict]:
    try:
        from pypdf import PdfReader
    except ImportError:
        return "", {"method": "pdf_missing_dependency", "pages": []}

    reader = PdfReader(str(path))
    pages: list[dict] = []
    text_parts: list[str] = []
    for index, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        pages.append({"page": index, "characters": len(page_text)})
        if page_text:
            text_parts.append(f"Page {index}\n{page_text}")
    text = "\n\n".join(text_parts)
    if text.strip():
        return text, {"method": "pypdf", "pages": pages}
    ocr_result = get_ocr_provider().extract(str(path))
    return ocr_result.text, {"method": "ocr", "provider": ocr_result.provider, "pages": ocr_result.pages, "confidence": ocr_result.confidence}
